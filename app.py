"""
Face Attendance System — Streamlit front-end
Run with:  streamlit run app.py
"""

import os
import sys
import queue
import threading
from datetime import datetime
from io import BytesIO

import av
import cv2
import numpy as np
import pandas as pd
import streamlit as st
from PIL import Image
from streamlit_webrtc import webrtc_streamer, WebRtcMode, RTCConfiguration

ROOT = os.path.dirname(os.path.abspath(__file__))
if ROOT not in sys.path:
    sys.path.insert(0, ROOT)

from core.database import (
    init_db, register_person, get_all_persons, delete_person,
    mark_attendance, get_attendance_report, get_today_count,
)
from core.face_engine import register_face, remove_face, recognize_faces

KNOWN_FACES_DIR = os.path.join(ROOT, "data", "known_faces")
os.makedirs(KNOWN_FACES_DIR, exist_ok=True)
init_db()

# ─────────────────────────────────────────────────────────────────────────────
# Page config — must be FIRST Streamlit call
# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="FaceAttend",
    page_icon="⬡",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─────────────────────────────────────────────────────────────────────────────
# Global CSS — dark industrial theme with accent grid lines
# ─────────────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Space+Mono:wght@400;700&family=DM+Sans:wght@300;400;500;600&display=swap');

/* ── Reset & root ── */
html, body, [class*="css"] { font-family: 'DM Sans', sans-serif; }

/* ── Page background ── */
.stApp {
    background: #0b0d14;
    background-image:
        linear-gradient(rgba(99,102,241,.04) 1px, transparent 1px),
        linear-gradient(90deg, rgba(99,102,241,.04) 1px, transparent 1px);
    background-size: 40px 40px;
}

/* ── Hide Streamlit chrome ── */
#MainMenu, footer, header { visibility: hidden; }
div[data-testid="stSidebarNav"] { display: none !important; }
div[data-testid="collapsedControl"] { display: none; }

/* ── Sidebar ── */
section[data-testid="stSidebar"] {
    background: #0f1119 !important;
    border-right: 1px solid #1e2235;
}
section[data-testid="stSidebar"] > div { padding-top: 0 !important; }

/* ── Sidebar logo band ── */
.sb-logo {
    background: linear-gradient(135deg, #6366f1 0%, #8b5cf6 100%);
    margin: 0 -1rem 1.5rem -1rem;
    padding: 1.4rem 1.5rem;
    font-family: 'Space Mono', monospace;
    font-size: 1.1rem;
    font-weight: 700;
    color: #fff;
    letter-spacing: .04em;
}
.sb-logo span { opacity: .6; font-size: .7rem; display: block; margin-top: 2px; font-weight: 400; }

/* ── Sidebar radio (nav buttons) ── */
div[data-testid="stSidebar"] .stRadio > div {
    gap: 4px !important;
    flex-direction: column;
}
div[data-testid="stSidebar"] .stRadio label {
    background: transparent;
    border: none;
    border-radius: 8px;
    padding: .6rem 1rem !important;
    font-size: .9rem;
    color: #8892a4;
    cursor: pointer;
    transition: all .15s;
    width: 100%;
    margin: 0 !important;
}
div[data-testid="stSidebar"] .stRadio label:hover {
    background: #1a1e2e;
    color: #c7d0e0;
}
div[data-testid="stSidebar"] .stRadio label[data-selected="true"],
div[data-testid="stSidebar"] .stRadio input:checked + div {
    background: #1a1e2e;
    color: #818cf8;
    border-left: 3px solid #6366f1;
}

/* ── Hide radio circles ── */
div[data-testid="stSidebar"] .stRadio input { display: none !important; }

/* ── Stat counter ── */
.stat-box {
    background: #13161f;
    border: 1px solid #1e2235;
    border-radius: 12px;
    padding: 1rem 1.25rem;
    margin-top: .5rem;
}
.stat-box .label { font-size: .75rem; color: #4b5568; text-transform: uppercase; letter-spacing: .08em; }
.stat-box .value { font-family: 'Space Mono', monospace; font-size: 2.4rem; color: #6ee7b7; line-height: 1.1; }

/* ── Page titles ── */
.page-title {
    font-family: 'Space Mono', monospace;
    font-size: 1.45rem;
    font-weight: 700;
    color: #e2e8f0;
    letter-spacing: -.01em;
    margin-bottom: .25rem;
}
.page-sub { font-size: .85rem; color: #4b5568; margin-bottom: 1.5rem; }

/* ── Cards ── */
.card {
    background: #13161f;
    border: 1px solid #1e2235;
    border-radius: 14px;
    padding: 1.25rem 1.5rem;
    margin-bottom: .75rem;
}
.card-accent { border-left: 3px solid #6366f1; }

/* ── Metric tiles ── */
.metric-row { display: flex; gap: 12px; margin-bottom: 1.5rem; }
.metric-tile {
    flex: 1;
    background: #13161f;
    border: 1px solid #1e2235;
    border-radius: 12px;
    padding: 1rem 1.25rem;
}
.metric-tile .mt-label { font-size: .72rem; color: #4b5568; text-transform: uppercase; letter-spacing: .08em; margin-bottom: .3rem; }
.metric-tile .mt-value { font-family: 'Space Mono', monospace; font-size: 1.8rem; color: #e2e8f0; line-height: 1; }
.metric-tile .mt-value.green { color: #6ee7b7; }
.metric-tile .mt-value.indigo { color: #818cf8; }

/* ── Detection log badges ── */
.det-log { display: flex; flex-direction: column; gap: 5px; }
.det-badge {
    display: flex; align-items: center; gap: 8px;
    background: #13161f; border: 1px solid #1e2235;
    border-radius: 8px; padding: 6px 10px;
    font-size: .8rem; color: #8892a4;
    animation: slideIn .2s ease;
}
@keyframes slideIn { from { opacity:0; transform: translateX(6px); } to { opacity:1; transform: none; } }
.det-badge .dot { width: 7px; height: 7px; border-radius: 50%; flex-shrink: 0; }
.dot-green { background: #6ee7b7; box-shadow: 0 0 6px #6ee7b7; }
.dot-red   { background: #f87171; box-shadow: 0 0 6px #f87171; }
.dot-gray  { background: #4b5568; }
.det-badge .det-name { color: #e2e8f0; font-weight: 500; }
.det-badge .det-time { margin-left: auto; font-family: 'Space Mono', monospace; font-size: .72rem; color: #4b5568; }

/* ── Person rows ── */
.person-row {
    display: flex; align-items: center; gap: 14px;
    background: #13161f; border: 1px solid #1e2235;
    border-radius: 12px; padding: .9rem 1.1rem;
    margin-bottom: .5rem;
    transition: border-color .15s;
}
.person-row:hover { border-color: #2a2f45; }
.avatar {
    width: 42px; height: 42px; border-radius: 50%;
    background: linear-gradient(135deg, #6366f1, #8b5cf6);
    display: flex; align-items: center; justify-content: center;
    font-family: 'Space Mono', monospace; font-weight: 700;
    font-size: .95rem; color: #fff; flex-shrink: 0;
}
.person-name { font-size: .95rem; font-weight: 600; color: #e2e8f0; }
.person-meta { font-size: .78rem; color: #4b5568; margin-top: 1px; }
.person-actions { margin-left: auto; }

/* ── Streamlit widget overrides ── */
div[data-testid="stTextInput"] input,
div[data-testid="stDateInput"] input {
    background: #0f1119 !important;
    border: 1px solid #1e2235 !important;
    border-radius: 8px !important;
    color: #e2e8f0 !important;
    font-family: 'DM Sans', sans-serif !important;
}
div[data-testid="stTextInput"] input:focus,
div[data-testid="stDateInput"] input:focus {
    border-color: #6366f1 !important;
    box-shadow: 0 0 0 2px rgba(99,102,241,.2) !important;
}
.stButton > button {
    background: #6366f1 !important;
    color: #fff !important;
    border: none !important;
    border-radius: 8px !important;
    font-family: 'DM Sans', sans-serif !important;
    font-weight: 500 !important;
    transition: all .15s !important;
}
.stButton > button:hover { background: #5254cc !important; transform: translateY(-1px); }
button[kind="secondary"] {
    background: #1a1e2e !important;
    color: #f87171 !important;
    border: 1px solid #2a1f2a !important;
}
button[kind="secondary"]:hover { background: #2a1a1a !important; }

/* ── File uploader ── */
div[data-testid="stFileUploader"] {
    background: #0f1119;
    border: 1px dashed #1e2235;
    border-radius: 12px;
    padding: .5rem;
}

/* ── DataFrame ── */
.stDataFrame { border-radius: 12px; overflow: hidden; }
.stDataFrame thead th {
    background: #13161f !important;
    color: #6366f1 !important;
    font-family: 'Space Mono', monospace !important;
    font-size: .75rem !important;
    text-transform: uppercase;
    letter-spacing: .06em;
}
.stDataFrame tbody tr:hover td { background: #1a1e2e !important; }

/* ── WebRTC component ── */
.stWebRtcVideo video {
    border-radius: 14px !important;
    border: 1px solid #1e2235 !important;
}
div[data-testid="stWebRtc"] {
    border-radius: 14px;
    overflow: hidden;
    border: 1px solid #1e2235;
    background: #0b0d14;
}

/* ── Divider ── */
hr { border-color: #1e2235 !important; }

/* ── Success / error ── */
div[data-testid="stAlert"] { border-radius: 10px !important; }
</style>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────────────────
# Session state init
# ─────────────────────────────────────────────────────────────────────────────
if "det_log" not in st.session_state:
    st.session_state.det_log = []


# ─────────────────────────────────────────────────────────────────────────────
# WebRTC — frame callback architecture (avoids Windows DirectShow conflicts)
#
# WHY THIS APPROACH:
#   The class-based VideoProcessorBase holds a reference to the processor object
#   across Streamlit reruns, which can keep a lock on the DirectShow camera device.
#   Using a module-level queue + a plain callback function avoids that.
#   The callback runs on the aiortc worker thread — never the main thread —
#   so face_recognition math never blocks the UI.
# ─────────────────────────────────────────────────────────────────────────────
_event_queue: queue.Queue = queue.Queue(maxsize=30)
_recently_marked: dict = {}
_mark_lock = threading.Lock()


def _video_frame_callback(frame: av.VideoFrame) -> av.VideoFrame:
    """
    Called by aiortc on its own worker thread for every incoming frame.
    Never touches Streamlit state or widgets directly.
    Pushes detection events into _event_queue for the main thread to read.
    """
    img = frame.to_ndarray(format="bgr24")

    try:
        results = recognize_faces(img)
    except Exception:
        results = []

    now = datetime.now()

    for r in results:
        top, right_x, bottom, left = r["location"]
        is_match = r["matched"]
        color = (110, 231, 183) if is_match else (248, 113, 113)   # BGR: mint / red

        # Bounding box
        cv2.rectangle(img, (left, top), (right_x, bottom), color, 2)

        # Label pill background
        label = f"{r['name']}  {r['confidence']}%"
        (lw, lh), _ = cv2.getTextSize(label, cv2.FONT_HERSHEY_DUPLEX, 0.55, 1)
        cv2.rectangle(img, (left, top - lh - 12), (left + lw + 10, top), color, cv2.FILLED)
        cv2.putText(img, label, (left + 5, top - 5),
                    cv2.FONT_HERSHEY_DUPLEX, 0.55, (15, 17, 25), 1, cv2.LINE_AA)

        # Attendance + event
        if is_match and r["person_id"] is not None:
            pid = r["person_id"]
            with _mark_lock:
                last = _recently_marked.get(pid)
                too_soon = last and (now - last).seconds < 10
            if not too_soon:
                marked = mark_attendance(pid, r["name"])
                with _mark_lock:
                    _recently_marked[pid] = now
                ev = {"name": r["name"], "matched": True,
                      "marked": marked, "time": now.strftime("%H:%M:%S")}
            else:
                ev = None
        else:
            ev = {"name": "Unknown", "matched": False,
                  "marked": False, "time": now.strftime("%H:%M:%S")}

        if ev:
            try:
                _event_queue.put_nowait(ev)
            except queue.Full:
                try:
                    _event_queue.get_nowait()
                    _event_queue.put_nowait(ev)
                except queue.Empty:
                    pass

    return av.VideoFrame.from_ndarray(img, format="bgr24")


# ─────────────────────────────────────────────────────────────────────────────
# Excel helper — writes to BytesIO so st.download_button works without disk I/O
# ─────────────────────────────────────────────────────────────────────────────
def _build_excel_bytes(date_filter=None):
    records = get_attendance_report(date_filter)
    persons  = get_all_persons()
    if not records:
        return None
    persons_map = {p["id"]: p["role"] for p in persons}
    df = pd.DataFrame(records)
    df["role"] = df["person_id"].map(persons_map).fillna("—")
    df = df[["date","time","person_name","role","status"]]
    df.columns = ["Date","Time","Name","Role","Status"]
    df = df.sort_values(["Date","Time"], ascending=[False, True])

    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Attendance")
        ws = writer.sheets["Attendance"]
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        hf = Font(bold=True, color="FFFFFF", size=11)
        hfill = PatternFill(start_color="3730A3", end_color="3730A3", fill_type="solid")
        thin  = Side(style="thin", color="1E2235")
        brd   = Border(left=thin, right=thin, top=thin, bottom=thin)
        pfill = PatternFill(start_color="ECFDF5", end_color="ECFDF5", fill_type="solid")
        for i in range(1, len(df.columns)+1):
            c = ws.cell(row=1, column=i)
            c.font = hf; c.fill = hfill
            c.alignment = Alignment(horizontal="center"); c.border = brd
        for w, i in zip([14,12,24,14,10], range(1,6)):
            ws.column_dimensions[get_column_letter(i)].width = w
        for row in ws.iter_rows(min_row=2, max_row=len(df)+1):
            for cell in row:
                cell.border = brd
                if cell.column == 5 and cell.value == "Present":
                    cell.fill = pfill
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# Sidebar
# ─────────────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("""
        <div class="sb-logo">
            ⬡ FaceAttend
            <span>Facial Recognition Attendance</span>
        </div>
    """, unsafe_allow_html=True)

    NAV = [
        "📷  Live Attendance",
        "➕  Register Person",
        "👥  Manage People",
        "📋  Attendance Log",
    ]
    page = st.radio("nav", NAV, label_visibility="collapsed")

    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown(f"""
        <div class="stat-box">
            <div class="label">Today's attendance</div>
            <div class="value">{get_today_count():02d}</div>
        </div>
    """, unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────────────────
# Helper to render page header
# ─────────────────────────────────────────────────────────────────────────────
def page_header(title, sub=""):
    st.markdown(f'<div class="page-title">{title}</div>', unsafe_allow_html=True)
    if sub:
        st.markdown(f'<div class="page-sub">{sub}</div>', unsafe_allow_html=True)


# ═════════════════════════════════════════════════════════════════════════════
# PAGE 1 — Live Attendance
# ═════════════════════════════════════════════════════════════════════════════
if page == NAV[0]:
    page_header(
        "📷 Live Attendance",
        "Camera streams via WebRTC · face recognition runs in background thread · video never freezes"
    )

    col_cam, col_panel = st.columns([3, 1], gap="medium")

    with col_cam:
        # ── WebRTC streamer ───────────────────────────────────────────────────
        # KEY CAMERA FIX:
        #   video_frame_callback= (function, not a class) avoids the
        #   VideoProcessorBase lifecycle issue that holds DirectShow devices.
        #
        #   video_html_attrs forces the browser to show the LOCAL preview
        #   immediately without waiting for the round-trip annotated stream —
        #   this is why the "small preview" worked: it was the browser's own
        #   local track. We make the main feed behave the same way.
        #
        #   desired_playing_state is NOT set, so the user controls start/stop.
        ctx = webrtc_streamer(
            key="fa-live",
            mode=WebRtcMode.SENDRECV,
            video_frame_callback=_video_frame_callback,
            media_stream_constraints={
                "video": {
                    "width":  {"ideal": 1280},
                    "height": {"ideal": 720},
                    "frameRate": {"ideal": 30, "max": 30},
                },
                "audio": False,
            },
            async_processing=True,          # callback runs on thread pool, not event loop
            rtc_configuration=RTCConfiguration(
                iceServers=[
                    {"urls": ["stun:stun.l.google.com:19302"]},
                    {"urls": ["stun:stun1.l.google.com:19302"]},
                ]
            ),
            translations={
                "start": "▶  Start Camera",
                "stop":  "⏹  Stop Camera",
                "select_device": "Choose Camera",
            },
        )

    with col_panel:
        # ── Drain event queue and update session log ──────────────────────────
        if ctx.state.playing:
            drained = 0
            while drained < 20:
                try:
                    ev = _event_queue.get_nowait()
                    if ev["matched"] and ev["marked"]:
                        badge = ("dot-green", ev["name"], f"✓ marked {ev['time']}")
                    elif ev["matched"]:
                        badge = ("dot-gray",  ev["name"], f"already marked")
                    else:
                        badge = ("dot-red",   "Unknown",  ev["time"])
                    st.session_state.det_log.insert(0, badge)
                    drained += 1
                except queue.Empty:
                    break
            st.session_state.det_log = st.session_state.det_log[:50]

        # ── Render log ────────────────────────────────────────────────────────
        st.markdown('<div style="font-size:.8rem;font-weight:600;color:#4b5568;'
                    'text-transform:uppercase;letter-spacing:.08em;margin-bottom:.6rem">'
                    'DETECTIONS</div>', unsafe_allow_html=True)

        if not st.session_state.det_log:
            st.markdown('<div style="color:#2a2f45;font-size:.85rem;padding:.5rem 0">'
                        'Start the camera to begin.</div>', unsafe_allow_html=True)
        else:
            items = "".join(
                f'<div class="det-badge">'
                f'  <div class="dot {dc}"></div>'
                f'  <span class="det-name">{nm}</span>'
                f'  <span class="det-time">{tm}</span>'
                f'</div>'
                for dc, nm, tm in st.session_state.det_log[:20]
            )
            st.markdown(f'<div class="det-log">{items}</div>', unsafe_allow_html=True)

        if ctx.state.playing:
            st.rerun()


# ═════════════════════════════════════════════════════════════════════════════
# PAGE 2 — Register Person
# ═════════════════════════════════════════════════════════════════════════════
elif page == NAV[1]:
    page_header("➕ Register Person", "Add a new face to the recognition database")

    col_form, col_prev = st.columns([3, 2], gap="large")

    with col_form:
        st.markdown('<div class="card card-accent">', unsafe_allow_html=True)

        name     = st.text_input("Full name", placeholder="e.g. Yara Mohamed")
        role     = st.text_input("Role", placeholder="Student / Staff / Doctor…", value="Student")
        uploaded = st.file_uploader(
            "Face photo — clear, frontal, well-lit",
            type=["jpg","jpeg","png","bmp"],
        )

        st.markdown("<br>", unsafe_allow_html=True)
        register_clicked = st.button("✔  Register Person", use_container_width=True)
        st.markdown('</div>', unsafe_allow_html=True)

        if register_clicked:
            if not name.strip():
                st.error("Please enter a name.")
            elif not uploaded:
                st.error("Please upload a photo.")
            else:
                ext  = os.path.splitext(uploaded.name)[1] or ".jpg"
                safe = name.strip().replace(" ", "_")
                dest = os.path.join(KNOWN_FACES_DIR, f"{safe}{ext}")
                with open(dest, "wb") as f:
                    f.write(uploaded.getbuffer())
                pid     = register_person(name.strip(), role.strip() or "Student", dest)
                success = register_face(pid, name.strip(), dest)
                if success:
                    st.success(f"✅ **{name.strip()}** registered successfully!")
                    st.balloons()
                else:
                    delete_person(pid)
                    if os.path.exists(dest):
                        os.remove(dest)
                    st.error("❌ No face detected. Use a clear, frontal photo with one face visible.")

        st.markdown('<div style="margin-top:1rem;font-size:.8rem;color:#2a2f45">'
                    'Tips · Well-lit · Frontal face · No glasses or masks · One person only'
                    '</div>', unsafe_allow_html=True)

    with col_prev:
        if uploaded:
            img = Image.open(uploaded)
            img.thumbnail((320, 360))
            st.markdown('<div style="border:1px solid #1e2235;border-radius:14px;overflow:hidden">',
                        unsafe_allow_html=True)
            st.image(img, use_container_width=True)
            st.markdown('</div>', unsafe_allow_html=True)
        else:
            st.markdown(
                '<div style="height:260px;background:#0f1119;border:1px dashed #1e2235;'
                'border-radius:14px;display:flex;align-items:center;justify-content:center;'
                'color:#2a2f45;font-size:.9rem">No photo selected</div>',
                unsafe_allow_html=True,
            )


# ═════════════════════════════════════════════════════════════════════════════
# PAGE 3 — Manage People
# ═════════════════════════════════════════════════════════════════════════════
elif page == NAV[2]:
    persons = get_all_persons()
    page_header(
        "👥 Registered People",
        f"{len(persons)} person{'s' if len(persons) != 1 else ''} in the system"
    )

    if not persons:
        st.markdown(
            '<div class="card" style="color:#4b5568;text-align:center;padding:2.5rem">'
            'No people registered yet.<br>Go to <b>Register Person</b> to add someone.'
            '</div>',
            unsafe_allow_html=True,
        )
    else:
        for p in persons:
            initials = "".join(w[0].upper() for w in p["name"].split()[:2])
            col_person, col_del = st.columns([10, 1])
            with col_person:
                st.markdown(f"""
                    <div class="person-row">
                        <div class="avatar">{initials}</div>
                        <div>
                            <div class="person-name">{p['name']}</div>
                            <div class="person-meta">{p['role']} &nbsp;·&nbsp; since {p['registered_at'][:10]}</div>
                        </div>
                    </div>
                """, unsafe_allow_html=True)
            with col_del:
                st.markdown("<div style='margin-top:.35rem'>", unsafe_allow_html=True)
                if st.button("🗑", key=f"d{p['id']}", help=f"Remove {p['name']}"):
                    delete_person(p["id"])
                    remove_face(p["id"])
                    st.success(f"Removed **{p['name']}**.")
                    st.rerun()
                st.markdown("</div>", unsafe_allow_html=True)


# ═════════════════════════════════════════════════════════════════════════════
# PAGE 4 — Attendance Log
# ═════════════════════════════════════════════════════════════════════════════
elif page == NAV[3]:
    page_header("📋 Attendance Log", "View and export daily attendance records")

    col_f1, col_f2, col_f3 = st.columns([2, 1, 3])
    with col_f1:
        date_val   = st.date_input("Filter date", value=datetime.now())
    with col_f2:
        st.markdown("<br>", unsafe_allow_html=True)
        use_filter = st.checkbox("Filter", value=True)

    active_filter = date_val.strftime("%Y-%m-%d") if use_filter else None
    records       = get_attendance_report(active_filter)
    all_persons   = get_all_persons()
    today_c       = get_today_count()

    # Metrics row
    st.markdown(f"""
        <div class="metric-row">
            <div class="metric-tile">
                <div class="mt-label">Records shown</div>
                <div class="mt-value indigo">{len(records)}</div>
            </div>
            <div class="metric-tile">
                <div class="mt-label">Today's total</div>
                <div class="mt-value green">{today_c:02d}</div>
            </div>
            <div class="metric-tile">
                <div class="mt-label">Registered people</div>
                <div class="mt-value">{len(all_persons)}</div>
            </div>
        </div>
    """, unsafe_allow_html=True)

    if not records:
        st.markdown(
            '<div class="card" style="color:#4b5568;text-align:center;padding:2rem">'
            'No records found for this filter.</div>',
            unsafe_allow_html=True,
        )
    else:
        df = pd.DataFrame(records)
        df = df[["date","time","person_name","status"]].rename(
            columns={"date":"Date","time":"Time","person_name":"Name","status":"Status"}
        )
        st.dataframe(df, use_container_width=True, hide_index=True, height=420)

        excel_bytes = _build_excel_bytes(active_filter)
        if excel_bytes:
            label = date_val.strftime("%Y-%m-%d") if use_filter else "all"
            st.download_button(
                label="⬇  Download Excel Report",
                data=excel_bytes,
                file_name=f"attendance_{label}_{datetime.now().strftime('%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )