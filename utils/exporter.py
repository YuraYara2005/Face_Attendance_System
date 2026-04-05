import pandas as pd
import os
from datetime import datetime
from core.database import get_attendance_report, get_all_persons


def export_to_excel(date_filter: str = None, output_dir: str = None) -> str:
    """
    Export attendance records to a formatted Excel file.
    Returns the path of the saved file.
    """
    records = get_attendance_report(date_filter)
    persons = get_all_persons()
    persons_map = {p['id']: p['role'] for p in persons}

    if not records:
        return None

    df = pd.DataFrame(records)
    df['role'] = df['person_id'].map(persons_map).fillna('—')
    df = df[['date', 'time', 'person_name', 'role', 'status']]
    df.columns = ['Date', 'Time', 'Name', 'Role', 'Status']
    df = df.sort_values(['Date', 'Time'], ascending=[False, True])

    if output_dir is None:
        output_dir = os.path.join(os.path.dirname(__file__), '..', 'data', 'attendance_logs')
    os.makedirs(output_dir, exist_ok=True)

    label = date_filter if date_filter else 'all'
    filename = f'attendance_{label}_{datetime.now().strftime("%H%M%S")}.xlsx'
    filepath = os.path.join(output_dir, filename)

    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Attendance')

        workbook = writer.book
        worksheet = writer.sheets['Attendance']

        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='2B4C7E', end_color='2B4C7E', fill_type='solid')
        thin = Side(style='thin', color='CCCCCC')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col_num, col_name in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border

        col_widths = [14, 12, 24, 14, 10]
        for i, width in enumerate(col_widths, 1):
            worksheet.column_dimensions[get_column_letter(i)].width = width

        present_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
        for row in worksheet.iter_rows(min_row=2, max_row=len(df) + 1):
            for cell in row:
                cell.border = border
                if cell.column == 5 and cell.value == 'Present':
                    cell.fill = present_fill

    return filepath
